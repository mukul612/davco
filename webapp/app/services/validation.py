"""
Upload validation (before processing) and output validation (before the
job is allowed to report "complete"). Both raise ValidationError with a
human-readable message on failure -- callers turn that into an HTTP 4xx or
a failed job status, never a raw traceback shown to the user.
"""
import hashlib
from pathlib import Path

import openpyxl
import fitz  # pymupdf

from app.config import MAX_UPLOAD_BYTES


class ValidationError(Exception):
    pass


def validate_upload(filename: str, content: bytes) -> None:
    if not filename.lower().endswith(".pdf"):
        raise ValidationError("Only .pdf files are accepted.")
    if not content:
        raise ValidationError("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        mb = MAX_UPLOAD_BYTES / (1024 * 1024)
        raise ValidationError(f"The uploaded file is larger than the {mb:.0f} MB limit.")
    if not content.startswith(b"%PDF-"):
        raise ValidationError("The uploaded file is not a valid PDF (missing PDF header).")
    try:
        doc = fitz.open(stream=content, filetype="pdf")
        page_count = len(doc)
        doc.close()
    except Exception as e:
        raise ValidationError(f"The uploaded file could not be read as a PDF: {e}") from e
    if page_count == 0:
        raise ValidationError("The uploaded PDF has no pages.")


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def validate_outputs(
    ballooned_pdf: Path,
    excel_path: Path,
    template_path: Path,
    template_hash_before: str,
    expected_min_rows: int = 1,
) -> list[str]:
    """Returns a list of warning strings on soft issues. Raises
    ValidationError on anything that means the job must be marked FAILED."""
    warnings: list[str] = []

    if not ballooned_pdf.is_file() or ballooned_pdf.stat().st_size == 0:
        raise ValidationError("Ballooned PDF was not created or is empty.")
    try:
        doc = fitz.open(ballooned_pdf)
        if len(doc) == 0:
            raise ValidationError("Ballooned PDF has no pages.")
        doc.close()
    except ValidationError:
        raise
    except Exception as e:
        raise ValidationError(f"Ballooned PDF could not be opened: {e}") from e

    if not excel_path.is_file() or excel_path.stat().st_size == 0:
        raise ValidationError("Completed Excel file was not created or is empty.")
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise ValidationError(f"Completed Excel file could not be opened: {e}") from e
    if "Sheet1" not in wb.sheetnames:
        raise ValidationError("Completed Excel file is missing the expected 'Sheet1' worksheet.")
    ws = wb["Sheet1"]

    populated_rows = 0
    for row in ws.iter_rows(min_row=20, max_row=min(ws.max_row, 90), min_col=1, max_col=2):
        item_cell, spec_cell = row
        if item_cell.value is not None and spec_cell.value not in (None, ""):
            populated_rows += 1
    if populated_rows < expected_min_rows:
        warnings.append(
            f"Only {populated_rows} dimensional-table row(s) were populated -- "
            f"double-check the drawing was read correctly."
        )

    if template_path.is_file():
        template_hash_after = file_hash(template_path)
        if template_hash_after != template_hash_before:
            # This should be structurally impossible (we only ever copy from
            # the template, never open it for writing) -- treat it as a hard
            # failure rather than silently serving a possibly-corrupted asset.
            raise ValidationError(
                "SAFETY CHECK FAILED: the master Excel template changed on disk "
                "during processing. Aborting -- this must be investigated."
            )

    return warnings
