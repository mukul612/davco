"""
Tests for the deterministic parts of the existing dimensional-layout
engine, exercised through the web app's layout_engine wrapper against the
REAL master templates (never modifying them). No AI/Claude calls happen
here -- callouts are hand-crafted fixtures, exactly like the skill would
produce, to verify the mechanical Excel-filling and PDF-ballooning logic
the web app relies on unchanged.
"""
import fitz
import openpyxl

from app.config import CUSTOMER_TEMPLATE, SUPPLIER_TEMPLATE
from app.layout_engine import draw_balloons, fill_excel
from app.services.validation import file_hash

SAMPLE_CALLOUTS = [
    {"item": 1, "page": 0, "x": 100, "y": 100, "destination": "table", "text": "3.50 [89.0] MAX"},
    {"item": 2, "page": 0, "x": 150, "y": 150, "destination": "table", "text": "APTIV CONNECTOR P/N 1; APTIV TERMINAL P/N 2"},
    {
        "item": 3, "page": 0, "x": 200, "y": 200,
        "destination": "customer_part_number", "value": "9999999", "customer_name": "VOLVO",
        "text": "VOLVO P/N: 9999999",
    },
    {"item": 4, "page": 0, "x": 250, "y": 250, "destination": "davco_part_number", "value": "TEST-001", "text": "PART NO: TEST-001"},
    {"item": 5, "page": 0, "x": 300, "y": 300, "destination": "revision", "value": "A", "text": "REVISION: A"},
    {"item": 6, "page": 0, "x": 350, "y": 350, "destination": "released_date", "value": "01/01/24", "text": "RELEASED: 01/01/24"},
]


def _sample_pdf(tmp_path):
    doc = fitz.open()
    doc.new_page(width=600, height=400)
    p = tmp_path / "source.pdf"
    doc.save(p)
    doc.close()
    return p


def test_draw_balloons_creates_new_pdf_without_touching_source(tmp_path):
    src = _sample_pdf(tmp_path)
    src_hash_before = file_hash(src)
    out_pdf = tmp_path / "ballooned.pdf"

    draw_balloons(str(src), SAMPLE_CALLOUTS, str(out_pdf), 15.0, None, 2.0)

    assert out_pdf.is_file() and out_pdf.stat().st_size > 0
    assert file_hash(src) == src_hash_before  # source untouched
    doc = fitz.open(out_pdf)
    assert len(doc) == 1
    doc.close()


def test_fill_excel_supplier_routes_customer_pn_to_table(tmp_path):
    """Supplier template has no Customer fields at all -- a
    customer_part_number callout must automatically fall back into the
    dimensional table rather than being lost."""
    template_hash_before = file_hash(SUPPLIER_TEMPLATE)
    out = tmp_path / "out_supplier.xlsx"

    _out_path, notes, rows_used, header_used = fill_excel(
        str(SUPPLIER_TEMPLATE), SAMPLE_CALLOUTS, str(out), "Supplier", None,
    )

    assert file_hash(SUPPLIER_TEMPLATE) == template_hash_before  # master untouched
    assert "customer_part_number" not in header_used
    assert rows_used == 3  # items 1, 2, and 3 (fallen back) -- 4/5/6 are header-bound
    wb = openpyxl.load_workbook(out)
    ws = wb["Sheet1"]
    item_numbers = [ws.cell(row=r, column=1).value for r in range(20, 30) if ws.cell(row=r, column=1).value]
    assert item_numbers == [1, 2, 3]


def test_fill_excel_customer_routes_header_fields_and_leaves_gaps(tmp_path):
    template_hash_before = file_hash(CUSTOMER_TEMPLATE)
    out = tmp_path / "out_customer.xlsx"

    _out_path, notes, rows_used, header_used = fill_excel(
        str(CUSTOMER_TEMPLATE), SAMPLE_CALLOUTS, str(out), "Customer", None,
    )

    assert file_hash(CUSTOMER_TEMPLATE) == template_hash_before  # master untouched
    assert rows_used == 2  # only items 1 and 2 stay in the table
    assert header_used["customer_part_number"] == "O5"
    assert header_used["customer_name"] == "O3"
    assert header_used["davco_part_number"] == "A7"
    assert header_used["revision"] == "J7"

    wb = openpyxl.load_workbook(out)
    ws = wb["Sheet1"]
    # Gap check: items 3-6 must NOT appear as table rows -- table stops at 2.
    item_numbers = [ws.cell(row=r, column=1).value for r in range(20, 30) if ws.cell(row=r, column=1).value]
    assert item_numbers == [1, 2]
    assert ws["O3"].value == "VOLVO"
    assert ws["O5"].value == "[Balloon 3] 9999999"
    assert ws["A7"].value == "[Balloon 4] TEST-001"
    assert ws["J7"].value == "[Balloon 5] A"
    assert "01/01/24" in ws["A10"].value  # released-date fallback field
