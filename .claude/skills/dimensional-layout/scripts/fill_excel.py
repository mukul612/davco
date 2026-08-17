"""
fill_excel.py

Copies a blank DIM Excel template (Supplier or Customer -- auto-detected
from the filename, or given explicitly with --layout-type) to a new file
and fills it in. The template is NEVER opened for writing / NEVER modified.

Every callout in callouts_json has a "destination": either "table" (a row
in the Specification / Dimension / Characteristic grid) or a header
metadata field ("davco_part_number", "revision", "released_date",
"customer_name", "customer_part_number"). This script:

  - writes table-destined callouts into the dimensional table, using the
    callout's OWN balloon/item number in column A (NOT a re-flowed 1,2,3
    sequence) -- so gaps are expected and correct wherever a balloon's
    destination is a header field instead of a table row.
  - writes header-destined callouts into the matching header cell for the
    selected template, prefixed with "[Balloon N]" so the reader can trace
    it back to the drawing. If the selected template has no field for a
    given destination (e.g. a Customer Part Number balloon while filling a
    Supplier layout, which has no Customer fields at all), that callout is
    automatically redirected into the table instead -- every balloon always
    ends up SOMEWHERE, per the "every balloon has exactly one destination"
    rule.
  - writes a "limits"/tolerance value into the separate Specification /
    Limits column when the template has one (the Customer template does;
    the Supplier template doesn't) and the callout provides one. Otherwise
    the full callout text stays together in one cell.
  - fills the non-ballooned Description/Part-Name field from --header if
    given (no balloon number attached -- it's just autofilled metadata).

All existing formatting, merged cells, fonts, borders, row heights and
column widths are preserved. Only cell values change, except when a part
has more table rows than the template pre-built -- those new rows copy the
style of the last template item row.

Usage:
    python fill_excel.py <template_xlsx> <callouts_json> <output_xlsx> [--layout-type Supplier|Customer] [--header header_json]

callouts_json shape (one list, shared with create_balloons.py -- extra
keys like x/y/leader_to are simply ignored here):
[
  {"item": 7, "destination": "table", "text": "6.58 [167.2] MAX"},
  {"item": 10, "destination": "table", "text": "APTIV CONNECTOR P/N 12015792; APTIV TERMINAL P/N 12124580; 12V-150W NOMINAL; NON-POLARIZED CONNECTOR"},
  {"item": 20, "destination": "customer_part_number", "value": "3594343C95", "customer_name": "NAVISTAR"},
  {"item": 23, "destination": "davco_part_number", "value": "382932NAV-25"},
  {"item": 24, "destination": "revision", "value": "D"},
  {"item": 25, "destination": "released_date", "value": "07/19/13"}
]

A callout may also carry "limits": "±.010 [.25]" to populate the Customer
template's separate Limits column.

header_json shape (all optional, none of these carry balloon numbers):
{"description": "FP382 ASM, FLUID HEAT W/ 12V PRE-HEAT"}
"""
import argparse
import copy
import json
import os
import shutil
import sys

import openpyxl

ITEM_HEADER_TEXT = "#"
SPEC_HEADER_TEXT = "Specification / Dimension / Characteristic"
LIMITS_HEADER_TEXT = "Limits"

# Header metadata cell maps, hand-verified against the two templates in
# "C:\Davco\Layout by AI". Each entry is the VALUE cell (not the label
# cell) that already exists in the blank template.
SUPPLIER_HEADER_CELLS = {
    "davco_part_number": "N7",
    "revision": "X7",
    "description": "N10",
    # Neither template has a dedicated "Released Date" field. Fall back to
    # the otherwise-unused "Lay-out No." box (A10 on both templates) with
    # an inline label so it's self-explanatory. Flag this in the review
    # log every time it's used.
    "released_date": "A10",
    # No Customer fields exist on the Supplier template at all -- omitted
    # on purpose so the fallback-to-table logic below kicks in.
}

CUSTOMER_HEADER_CELLS = {
    "customer_name": "O3",
    "customer_part_number": "O5",
    "davco_part_number": "A7",
    "revision": "J7",
    "description": "O9",
    "released_date": "A10",  # same fallback reasoning as the Supplier template
}

TEMPLATE_MAPS = {
    "supplier": SUPPLIER_HEADER_CELLS,
    "customer": CUSTOMER_HEADER_CELLS,
}

RELEASED_DATE_FALLBACK_NOTE = (
    "released_date has no dedicated header field in this template -- "
    "written into the 'Lay-out No.' box (A10) instead, prefixed so it's "
    "self-explanatory. Flag for the user in the review log."
)


def detect_layout_type(template_path, explicit=None):
    if explicit:
        key = explicit.strip().lower()
        if key in TEMPLATE_MAPS:
            return key
        raise ValueError(f"--layout-type must be 'Supplier' or 'Customer', got {explicit!r}")
    name = os.path.basename(template_path).lower()
    if "customer" in name:
        return "customer"
    if "supplier" in name:
        return "supplier"
    raise ValueError(
        f"Could not tell whether {template_path!r} is the Supplier or Customer "
        f"template from its filename -- pass --layout-type explicitly."
    )


def find_columns(ws):
    """Locate the header row and the Item# / Specification / (optional)
    Limits columns by scanning for their known header text, rather than
    hardcoding cell addresses -- keeps this working if a template layout
    ever shifts."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == SPEC_HEADER_TEXT:
                spec_col = cell.column
                header_row = cell.row
                item_col = None
                for c2 in ws[header_row]:
                    if isinstance(c2.value, str) and c2.value.strip() == ITEM_HEADER_TEXT:
                        item_col = c2.column
                        break
                if item_col is None:
                    item_col = spec_col - 1

                limits_col = None
                for r in (header_row - 1, header_row):
                    for c2 in ws[r]:
                        if isinstance(c2.value, str) and c2.value.strip().rstrip("/").strip() == LIMITS_HEADER_TEXT:
                            limits_col = c2.column
                            break
                    if limits_col:
                        break

                return header_row, item_col, spec_col, limits_col
    raise RuntimeError(f"Could not find header cell with text {SPEC_HEADER_TEXT!r} in template")


def find_item_rows(ws, header_row, item_col):
    """Return the sorted list of template item-row numbers (rows below the
    header that the template pre-built for table entries, whether or not
    they still carry the template's own placeholder number)."""
    rows = []
    r = header_row + 1
    blank_streak = 0
    while blank_streak < 3 and r < header_row + 500:
        v = ws.cell(row=r, column=item_col).value
        if isinstance(v, (int, float)):
            rows.append(r)
            blank_streak = 0
        else:
            blank_streak += 1
        r += 1
    return rows


def copy_row_style(ws, src_row, dst_row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy.copy(src_cell.protection)
            dst_cell.alignment = copy.copy(src_cell.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == src_row and mr.max_row == src_row:
            ws.merge_cells(start_row=dst_row, start_column=mr.min_col,
                            end_row=dst_row, end_column=mr.max_col)


def fill(template_path, callouts, output_path, layout_type=None, header=None):
    layout_key = detect_layout_type(template_path, layout_type)
    header_cells = TEMPLATE_MAPS[layout_key]

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    header_row, item_col, spec_col, limits_col = find_columns(ws)
    template_rows = find_item_rows(ws, header_row, item_col)

    notes = []
    table_entries = []
    for entry in callouts:
        dest = entry.get("destination", "table")
        if dest != "table" and dest not in header_cells:
            notes.append(
                f"item {entry['item']}: destination {dest!r} has no field in the "
                f"{layout_key} template -- written into the dimensional table instead."
            )
            dest = "table"
        entry = dict(entry, destination=dest)
        if dest == "table":
            table_entries.append(entry)

    table_entries.sort(key=lambda e: e["item"])

    # Clear every pre-filled template item number first -- gaps are
    # intentional, so we don't want stray leftover numbers on rows that
    # end up unused.
    for r in template_rows:
        ws.cell(row=r, column=item_col).value = None

    max_col = ws.max_column
    last_row = template_rows[-1] if template_rows else header_row
    rows_used = 0
    for i, entry in enumerate(table_entries):
        if i < len(template_rows):
            target_row = template_rows[i]
        else:
            target_row = last_row + 1
            copy_row_style(ws, last_row, target_row, 1, max_col)
            last_row = target_row
        rows_used += 1

        ws.cell(row=target_row, column=item_col).value = entry["item"]
        ws.cell(row=target_row, column=spec_col).value = entry.get("text", entry.get("value", ""))
        if limits_col and entry.get("limits"):
            ws.cell(row=target_row, column=limits_col).value = entry["limits"]

    header_used = {}
    for entry in callouts:
        dest = entry.get("destination", "table")
        if dest == "table" or dest not in header_cells:
            continue
        coord = header_cells[dest]
        value = entry.get("value", entry.get("text", ""))
        if dest == "customer_name":
            cell_text = value
        else:
            cell_text = f"[Balloon {entry['item']}] {value}"
            if dest == "released_date" and coord == "A10":
                cell_text = f"[Balloon {entry['item']}] Released: {value}"
                notes.append(f"item {entry['item']} (released_date): {RELEASED_DATE_FALLBACK_NOTE}")
        ws[coord] = cell_text
        header_used[dest] = coord

        # The customer name itself isn't a separate balloon -- it rides
        # along on the customer_part_number callout (rule: "the customer
        # name does not need to be ballooned unless specifically
        # required"). Auto-fill it here, plain, with no balloon prefix.
        if dest == "customer_part_number" and entry.get("customer_name") and "customer_name" in header_cells:
            ws[header_cells["customer_name"]] = entry["customer_name"]
            header_used["customer_name"] = header_cells["customer_name"]

    if header and header.get("description"):
        desc_coord = header_cells.get("description")
        if desc_coord:
            ws[desc_coord] = header["description"]

    wb.save(output_path)
    print(f"Excel {layout_key} layout written to: {output_path}")
    print(f"  header row: {header_row}, item col: {item_col}, spec col: {spec_col}, limits col: {limits_col}")
    print(f"  table rows filled: {rows_used} (item numbers: {[e['item'] for e in table_entries]})")
    print(f"  header fields filled: {header_used}")
    for n in notes:
        print(f"  NOTE: {n}")
    return output_path, notes, rows_used, header_used


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("template_xlsx")
    ap.add_argument("callouts_json")
    ap.add_argument("output_xlsx")
    ap.add_argument("--layout-type", choices=["Supplier", "Customer"], default=None)
    ap.add_argument("--header", default=None, help="path to header_json (non-ballooned metadata like Description)")
    args = ap.parse_args()

    if not os.path.isfile(args.template_xlsx):
        print(f"ERROR: template not found: {args.template_xlsx}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(args.callouts_json):
        print(f"ERROR: callouts json not found: {args.callouts_json}", file=sys.stderr)
        sys.exit(1)

    with open(args.callouts_json, "r", encoding="utf-8-sig") as f:
        callouts = json.load(f)

    header = None
    if args.header and os.path.isfile(args.header):
        with open(args.header, "r", encoding="utf-8-sig") as f:
            header = json.load(f)

    fill(args.template_xlsx, callouts, args.output_xlsx, args.layout_type, header)
