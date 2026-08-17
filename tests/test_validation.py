import fitz
import openpyxl
import pytest

from app.config import SUPPLIER_TEMPLATE
from app.services.validation import ValidationError, validate_outputs, validate_upload


def test_valid_pdf_accepted(sample_pdf_bytes):
    validate_upload("drawing.pdf", sample_pdf_bytes)  # should not raise


def test_non_pdf_extension_rejected(sample_pdf_bytes):
    with pytest.raises(ValidationError, match="Only .pdf"):
        validate_upload("drawing.docx", sample_pdf_bytes)


def test_empty_file_rejected(empty_pdf_bytes):
    with pytest.raises(ValidationError, match="empty"):
        validate_upload("drawing.pdf", empty_pdf_bytes)


def test_non_pdf_content_rejected(not_a_pdf_bytes):
    with pytest.raises(ValidationError, match="not a valid PDF"):
        validate_upload("drawing.pdf", not_a_pdf_bytes)


def test_oversized_upload_rejected(sample_pdf_bytes, monkeypatch):
    import app.services.validation as v
    monkeypatch.setattr(v, "MAX_UPLOAD_BYTES", 10)  # smaller than our sample
    with pytest.raises(ValidationError, match="larger than"):
        validate_upload("drawing.pdf", sample_pdf_bytes)


def test_validate_outputs_missing_pdf(tmp_path):
    excel = tmp_path / "out.xlsx"
    excel.write_bytes(b"not real but non-empty")
    with pytest.raises(ValidationError, match="Ballooned PDF"):
        validate_outputs(tmp_path / "missing.pdf", excel, SUPPLIER_TEMPLATE, "irrelevant")


def test_validate_outputs_empty_pdf(tmp_path):
    pdf = tmp_path / "empty.pdf"
    pdf.write_bytes(b"")
    excel = tmp_path / "out.xlsx"
    excel.write_bytes(b"x")
    with pytest.raises(ValidationError, match="Ballooned PDF"):
        validate_outputs(pdf, excel, SUPPLIER_TEMPLATE, "irrelevant")


def test_validate_outputs_missing_excel(tmp_path):
    doc = fitz.open()
    doc.new_page()
    pdf = tmp_path / "ok.pdf"
    doc.save(pdf)
    with pytest.raises(ValidationError, match="Excel"):
        validate_outputs(pdf, tmp_path / "missing.xlsx", SUPPLIER_TEMPLATE, "irrelevant")


def test_validate_outputs_good_files_pass(tmp_path):
    doc = fitz.open()
    doc.new_page()
    pdf = tmp_path / "ok.pdf"
    doc.save(pdf)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=20, column=1).value = 1
    ws.cell(row=20, column=2).value = "3.50 [89.0] MAX"
    excel = tmp_path / "ok.xlsx"
    wb.save(excel)

    from app.services.validation import file_hash
    warnings = validate_outputs(
        pdf, excel, SUPPLIER_TEMPLATE, file_hash(SUPPLIER_TEMPLATE), expected_min_rows=1,
    )
    assert warnings == []


def test_validate_outputs_flags_template_mutation(tmp_path):
    """If the master template's hash changed during processing, that's a
    hard failure -- it should never happen since we only ever copy from
    it, but if it does, we must not silently serve a possibly-corrupted
    template."""
    doc = fitz.open()
    doc.new_page()
    pdf = tmp_path / "ok.pdf"
    doc.save(pdf)
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    excel = tmp_path / "ok.xlsx"
    wb.save(excel)

    with pytest.raises(ValidationError, match="SAFETY CHECK FAILED"):
        validate_outputs(pdf, excel, SUPPLIER_TEMPLATE, "0" * 64)
